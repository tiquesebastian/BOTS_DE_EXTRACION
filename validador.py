"""
OCR Pro - Extractor Inteligente de Campos (Versión Definitiva)
Soporta: PDF, TIF, TIFF, JPG, PNG
Detecta automáticamente patrones "CLAVE : VALOR"
Extrae los valores exactos que buscas
"""

import aspose.ocr as ocr
import os
import re
import tempfile
import time
import threading
from datetime import datetime
from tkinter import *
from tkinter import ttk, filedialog, messagebox, scrolledtext
from tkinter import font as tkfont
from pdf2image import convert_from_path
from PIL import Image, ImageEnhance, ImageFilter, ImageOps
import PIL.Image
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import numpy as np
from difflib import get_close_matches

# ============================================================
# CONFIGURACIÓN
# ============================================================

DPI_POR_DEFECTO = 300
MEJORAR_IMAGEN = True
SIMILITUD_MINIMA = 0.7      # Para fuzzy matching
LONGITUD_VALOR = 300        # Caracteres a extraer después de la palabra

# Colores
COLOR_PRIMARIO = "#2c3e50"
COLOR_SECUNDARIO = "#3498db"
COLOR_EXITO = "#27ae60"
COLOR_ERROR = "#e74c3c"
COLOR_ADVERTENCIA = "#f39c12"
COLOR_FONDO = "#ecf0f1"
COLOR_TEXTO = "#2c3e50"

# --- POPPLER ---
POPPLER_PATH = r'C:\poppler-25.11.0\Library\bin'
if not os.path.exists(POPPLER_PATH):
    posibles = [
        r'C:\poppler-25.11.0\Library\bin',
        r'C:\poppler-25.11.0\bin',
        r'C:\poppler\Library\bin',
        r'C:\poppler\bin',
    ]
    for ruta in posibles:
        if os.path.exists(ruta):
            POPPLER_PATH = ruta
            break
    else:
        messagebox.showwarning("Poppler", "No se encontró Poppler. Los PDF no se podrán procesar.")

print(f"Poppler en: {POPPLER_PATH}")

# ============================================================
# CLASE PRINCIPAL
# ============================================================

class AppOCR:
    def __init__(self, root):
        self.root = root
        self.root.title("OCR Pro - Extractor Inteligente de Campos")
        self.root.geometry("1500x850")
        self.root.configure(bg=COLOR_FONDO)
        
        self.archivos_seleccionados = []
        self.resultados = []
        self.procesando = False
        self.cancelar = False
        self.api = None
        self.carpeta_raiz = ""
        self.filtro_nombre = ""
        
        # Fuentes
        fuente_titulo = tkfont.Font(family="Segoe UI", size=16, weight="bold")
        fuente_normal = tkfont.Font(family="Segoe UI", size=10)
        fuente_boton = tkfont.Font(family="Segoe UI", size=10, weight="bold")
        
        self.fuente_titulo = fuente_titulo
        self.fuente_normal = fuente_normal
        self.fuente_boton = fuente_boton
        
        self.crear_interfaz()
        self.inicializar_ocr()
        self.log("✅ Aplicación iniciada - Extracción inteligente de campos", "exito")
    
    # ------------------------------------------------------------
    # INTERFAZ GRÁFICA
    # ------------------------------------------------------------
    def crear_interfaz(self):
        main_frame = Frame(self.root, bg=COLOR_FONDO)
        main_frame.pack(fill=BOTH, expand=True, padx=10, pady=10)
        
        # Barra superior
        barra = Frame(main_frame, bg=COLOR_PRIMARIO, height=60)
        barra.pack(fill=X, pady=(0,10))
        barra.pack_propagate(False)
        Label(barra, text="🔍 OCR Pro - Extractor Inteligente de Campos",
              font=self.fuente_titulo, bg=COLOR_PRIMARIO, fg="white").pack(side=LEFT, padx=20, pady=10)
        
        btn_test = Button(barra, text="🧪 Probar OCR", font=self.fuente_boton,
                          bg=COLOR_SECUNDARIO, fg="white", command=self.probar_ocr_manual)
        btn_test.pack(side=RIGHT, padx=10, pady=10)
        
        # Panel izquierdo
        panel_izq = Frame(main_frame, bg=COLOR_FONDO, width=540)
        panel_izq.pack(side=LEFT, fill=Y, padx=(0,10))
        panel_izq.pack_propagate(False)
        
        # Filtro por nombre
        g_filtro = LabelFrame(panel_izq, text="🔍 FILTRAR POR NOMBRE (opcional)",
                              font=self.fuente_normal, bg=COLOR_FONDO, fg=COLOR_PRIMARIO)
        g_filtro.pack(fill=X, pady=5, padx=5)
        Label(g_filtro, text="Ej: 'OTROS', 'FACTURA', 'RESULTADO' (vacío = todos)",
              font=("Segoe UI",9), bg=COLOR_FONDO, fg=COLOR_SECUNDARIO).pack(padx=10, pady=(5,0))
        self.entry_filtro = Entry(g_filtro, font=self.fuente_normal, bg="white")
        self.entry_filtro.pack(fill=X, padx=10, pady=5)
        self.lbl_filtro_activo = Label(g_filtro, text="", font=("Segoe UI",9,"italic"),
                                       bg=COLOR_FONDO, fg=COLOR_EXITO)
        self.lbl_filtro_activo.pack(padx=10, pady=(0,5))
        
        # Radicados específicos
        g_rad = LabelFrame(panel_izq, text="🔢 BUSCAR POR RADICADO (opcional)",
                           font=self.fuente_normal, bg=COLOR_FONDO, fg=COLOR_PRIMARIO)
        g_rad.pack(fill=X, pady=5, padx=5)
        Label(g_rad, text="Cualquier formato: espacios, comas, rangos (ej: 246016800-246016810)",
              font=("Segoe UI",9), bg=COLOR_FONDO, fg=COLOR_SECUNDARIO).pack(padx=10, pady=(5,0))
        self.text_radicados = Text(g_rad, height=3, font=self.fuente_normal)
        self.text_radicados.pack(fill=X, padx=10, pady=5)
        
        # Selección
        g_sel = LabelFrame(panel_izq, text="📁 SELECCIÓN", font=self.fuente_normal,
                           bg=COLOR_FONDO, fg=COLOR_PRIMARIO)
        g_sel.pack(fill=X, pady=5, padx=5)
        Button(g_sel, text="📂 Seleccionar Carpeta y Buscar", font=self.fuente_boton,
               bg=COLOR_SECUNDARIO, fg="white", command=self.buscar_archivos, height=2
               ).pack(fill=X, padx=10, pady=5)
        Button(g_sel, text="🗑️ Limpiar Lista", font=self.fuente_boton,
               bg=COLOR_ERROR, fg="white", command=self.limpiar_lista, height=2
               ).pack(fill=X, padx=10, pady=5)
        
        # ========== PALABRAS CLAVE ==========
        g_palabras = LabelFrame(panel_izq, text="🔑 PALABRAS CLAVE A BUSCAR (separadas por comas)",
                                font=self.fuente_normal, bg=COLOR_FONDO, fg=COLOR_PRIMARIO)
        g_palabras.pack(fill=X, pady=5, padx=5)
        Label(g_palabras, text="Ejemplo: paciente, fecha expedicion, documento, total",
              font=("Segoe UI",9), bg=COLOR_FONDO, fg=COLOR_SECUNDARIO).pack(padx=10, pady=(5,0))
        self.entry_mis_palabras = Entry(g_palabras, font=self.fuente_normal, bg="white")
        self.entry_mis_palabras.pack(fill=X, padx=10, pady=5)
        self.entry_mis_palabras.insert(0, "paciente, fecha expedicion, documento, fecha vencimiento, entidad, fecha ingreso, fecha egreso, total")
        
        # Tipo de búsqueda
        self.tipo_busqueda = StringVar(value="cualquiera")
        frame_tipo = Frame(g_palabras, bg=COLOR_FONDO)
        frame_tipo.pack(fill=X, padx=10, pady=5)
        Radiobutton(frame_tipo, text="CUALQUIER palabra (OR)", variable=self.tipo_busqueda,
                    value="cualquiera", bg=COLOR_FONDO).pack(side=LEFT, padx=5)
        Radiobutton(frame_tipo, text="TODAS las palabras (AND)", variable=self.tipo_busqueda,
                    value="todas", bg=COLOR_FONDO).pack(side=LEFT, padx=5)
        
        # Fuzzy matching opcional
        self.fuzzy_var = BooleanVar(value=True)
        Checkbutton(g_palabras, text="Usar coincidencia aproximada (fuzzy) para errores de OCR",
                    variable=self.fuzzy_var, bg=COLOR_FONDO).pack(anchor=W, padx=10, pady=5)
        
        # Configuración OCR
        g_cfg = LabelFrame(panel_izq, text="⚙️ CONFIGURACIÓN OCR",
                           font=self.fuente_normal, bg=COLOR_FONDO, fg=COLOR_PRIMARIO)
        g_cfg.pack(fill=X, pady=5, padx=5)
        
        frame_dpi = Frame(g_cfg, bg=COLOR_FONDO)
        frame_dpi.pack(fill=X, padx=10, pady=5)
        Label(frame_dpi, text="DPI (calidad):", font=self.fuente_boton,
              bg=COLOR_FONDO, fg=COLOR_SECUNDARIO).pack(side=LEFT)
        self.combo_dpi = ttk.Combobox(frame_dpi, values=["200","250","300","400"], width=8)
        self.combo_dpi.set(str(DPI_POR_DEFECTO))
        self.combo_dpi.pack(side=RIGHT)
        
        self.mejorar_var = BooleanVar(value=MEJORAR_IMAGEN)
        Checkbutton(g_cfg, text="Mejorar calidad de imagen (recomendado)",
                    variable=self.mejorar_var, bg=COLOR_FONDO).pack(anchor=W, padx=10, pady=5)
        
        frame_long = Frame(g_cfg, bg=COLOR_FONDO)
        frame_long.pack(fill=X, padx=10, pady=5)
        Label(frame_long, text="Caracteres a extraer:", font=self.fuente_normal,
              bg=COLOR_FONDO).pack(side=LEFT)
        self.entry_longitud = Entry(frame_long, width=10)
        self.entry_longitud.insert(0, str(LONGITUD_VALOR))
        self.entry_longitud.pack(side=RIGHT)
        
        Label(g_cfg, text="Formatos: PDF, TIF, TIFF, JPG, PNG",
              font=("Segoe UI",9), bg=COLOR_FONDO, fg=COLOR_EXITO).pack(padx=10, pady=5)
        
        # Acciones
        g_acc = LabelFrame(panel_izq, text="🚀 ACCIONES", font=self.fuente_normal,
                           bg=COLOR_FONDO, fg=COLOR_PRIMARIO)
        g_acc.pack(fill=X, pady=5, padx=5)
        self.btn_procesar = Button(g_acc, text="▶️ EXTRAER CAMPOS", font=self.fuente_boton,
                                   bg=COLOR_EXITO, fg="white", command=self.procesar_archivos, height=2)
        self.btn_procesar.pack(fill=X, padx=10, pady=5)
        self.btn_cancelar = Button(g_acc, text="⏹️ CANCELAR", font=self.fuente_boton,
                                   bg=COLOR_ERROR, fg="white", command=self.cancelar_proceso,
                                   state=DISABLED, height=2)
        self.btn_cancelar.pack(fill=X, padx=10, pady=5)
        self.btn_exportar = Button(g_acc, text="📊 Exportar a Excel", font=self.fuente_boton,
                                   bg=COLOR_SECUNDARIO, fg="white", command=self.exportar_excel,
                                   state=DISABLED, height=2)
        self.btn_exportar.pack(fill=X, padx=10, pady=5)
        
        # Progreso
        self.progress = ttk.Progressbar(panel_izq, mode='determinate')
        self.progress.pack(fill=X, pady=10, padx=5)
        self.lbl_progreso = Label(panel_izq, text="Listo", font=("Segoe UI",9),
                                  bg=COLOR_FONDO, fg=COLOR_TEXTO)
        self.lbl_progreso.pack()
        self.lbl_tiempo = Label(panel_izq, text="", font=("Segoe UI",9),
                                bg=COLOR_FONDO, fg=COLOR_SECUNDARIO)
        self.lbl_tiempo.pack()
        
        # Panel derecho
        panel_der = Frame(main_frame, bg=COLOR_FONDO)
        panel_der.pack(side=RIGHT, fill=BOTH, expand=True)
        
        # Lista de archivos
        g_lista = LabelFrame(panel_der, text="📋 ARCHIVOS SELECCIONADOS",
                             font=self.fuente_normal, bg=COLOR_FONDO, fg=COLOR_PRIMARIO)
        g_lista.pack(fill=BOTH, expand=True, pady=5, padx=5)
        frame_tree = Frame(g_lista, bg=COLOR_FONDO)
        frame_tree.pack(fill=BOTH, expand=True, padx=5, pady=5)
        scroll_y = Scrollbar(frame_tree)
        scroll_y.pack(side=RIGHT, fill=Y)
        scroll_x = Scrollbar(frame_tree, orient=HORIZONTAL)
        scroll_x.pack(side=BOTTOM, fill=X)
        self.tree = ttk.Treeview(frame_tree, columns=("radicado","archivo","tipo","paginas","tamaño","ruta"),
                                 yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)
        self.tree.heading("#0", text="N°")
        self.tree.heading("radicado", text="Radicado")
        self.tree.heading("archivo", text="Archivo")
        self.tree.heading("tipo", text="Tipo")
        self.tree.heading("paginas", text="Págs")
        self.tree.heading("tamaño", text="KB")
        self.tree.heading("ruta", text="Ruta")
        self.tree.column("#0", width=40)
        self.tree.column("radicado", width=100)
        self.tree.column("archivo", width=250)
        self.tree.column("tipo", width=50)
        self.tree.column("paginas", width=50)
        self.tree.column("tamaño", width=60)
        self.tree.column("ruta", width=450)
        self.tree.pack(fill=BOTH, expand=True)
        scroll_y.config(command=self.tree.yview)
        scroll_x.config(command=self.tree.xview)
        
        # Log
        g_log = LabelFrame(panel_der, text="📝 REGISTRO DE ACTIVIDAD",
                           font=self.fuente_normal, bg=COLOR_FONDO, fg=COLOR_PRIMARIO)
        g_log.pack(fill=BOTH, expand=True, pady=5, padx=5)
        self.log_text = scrolledtext.ScrolledText(g_log, height=10, font=("Consolas",9),
                                                  bg="#1e1e1e", fg="#d4d4d4")
        self.log_text.pack(fill=BOTH, expand=True, padx=5, pady=5)
        self.log_text.tag_config("exito", foreground="#4ec9b0")
        self.log_text.tag_config("error", foreground="#f48771")
        self.log_text.tag_config("info", foreground="#569cd6")
        self.log_text.tag_config("advertencia", foreground="#dcdcaa")
    
    # ------------------------------------------------------------
    # LOG
    # ------------------------------------------------------------
    def log(self, mensaje, tipo="info"):
        try:
            ts = datetime.now().strftime("%H:%M:%S")
            self.log_text.insert(END, f"[{ts}] {mensaje}\n", tipo)
            self.log_text.see(END)
            self.root.update()
        except:
            pass
    
    # ------------------------------------------------------------
    # OCR
    # ------------------------------------------------------------
    def inicializar_ocr(self):
        try:
            self.api = ocr.AsposeOcr()
            self.log("✅ Aspose OCR inicializado", "exito")
        except Exception as e:
            self.log(f"❌ Error OCR: {e}", "error")
            messagebox.showerror("OCR", "No se pudo inicializar el OCR")
    
    # ------------------------------------------------------------
    # NORMALIZACIÓN Y EXTRACCIÓN MEJORADAS
    # ------------------------------------------------------------
    def normalizar_texto(self, texto):
        """Limpia y normaliza el texto para facilitar la detección"""
        if not texto:
            return ""
        # Convertir a minúsculas
        texto = texto.lower()
        # Reemplazar tildes
        reemplazos_tildes = {
            'á': 'a', 'é': 'e', 'í': 'i', 'ó': 'o', 'ú': 'u',
            'ñ': 'n', 'ü': 'u', 'Á': 'a', 'É': 'e', 'Í': 'i', 'Ó': 'o', 'Ú': 'u'
        }
        for k, v in reemplazos_tildes.items():
            texto = texto.replace(k, v)
        # Reemplazar caracteres comunes mal reconocidos
        reemplazos_caracteres = {
            '0': 'o', '1': 'i', '3': 'e', '5': 's', '8': 'b',
            '@': 'a', '¢': 'c', '$': 's', '%': 'e', '&': 'y',
            '|': 'i', '!': 'i', '?': 'e', ';': ' ', ':': ' '
        }
        for k, v in reemplazos_caracteres.items():
            texto = texto.replace(k, v)
        # Eliminar caracteres no alfanuméricos ni espacios
        texto = re.sub(r'[^\w\s\.\,\-\:]', ' ', texto)
        # Unir palabras separadas por espacios múltiples
        texto = re.sub(r'\s+', ' ', texto).strip()
        return texto
    
    def extraer_valor(self, texto_norm, palabra, longitud, fuzzy=False):
        """
        Busca una palabra en el texto normalizado y extrae el valor siguiente.
        Soporta patrones como "CLAVE : VALOR" o "CLAVE VALOR"
        """
        palabra_limpia = re.sub(r'\s+', ' ', palabra.lower().strip())
        
        # Patrón 1: CLAVE : VALOR (con dos puntos y espacios)
        patron1 = rf'{palabra_limpia}\s*:\s*(.+?)(?:\n|$|\.|,|;)'
        match = re.search(patron1, texto_norm, re.IGNORECASE)
        if match:
            valor = match.group(1).strip()
            return valor[:longitud] if longitud > 0 else valor
        
        # Patrón 2: CLAVE VALOR (sin dos puntos, pero con espacio)
        patron2 = rf'{palabra_limpia}\s+([^\n\.\;,]+)'
        match = re.search(patron2, texto_norm, re.IGNORECASE)
        if match:
            valor = match.group(1).strip()
            return valor[:longitud] if longitud > 0 else valor
        
        # Patrón 3: Búsqueda simple (posición después de la palabra)
        pos = texto_norm.find(palabra_limpia)
        if pos != -1:
            resto = texto_norm[pos + len(palabra_limpia):].lstrip(' :;-\t\n')
            fin = min(len(resto), longitud)
            for delim in ['.', ',', '\n', ';']:
                idx = resto.find(delim)
                if idx != -1 and idx < fin:
                    fin = idx
            return resto[:fin].strip()
        
        # Búsqueda difusa (si está activada)
        if fuzzy:
            palabras_texto = texto_norm.split()
            for i, w in enumerate(palabras_texto):
                coinc = get_close_matches(w, [palabra_limpia], n=1, cutoff=SIMILITUD_MINIMA)
                if coinc:
                    resto = ' '.join(palabras_texto[i+1:])
                    valor = resto[:longitud].strip()
                    for delim in ['.', ',', '\n', ';']:
                        idx = valor.find(delim)
                        if idx != -1:
                            valor = valor[:idx]
                    return valor
        
        return None
    
    # ------------------------------------------------------------
    # PROCESAMIENTO DE ARCHIVOS (OCR)
    # ------------------------------------------------------------
    def parsear_radicados(self, texto):
        if not texto or not texto.strip():
            return None
        texto_norm = re.sub(r'\s+', ' ', texto.strip())
        radicados = set()
        rangos = re.findall(r'(\d{8,10})\s*-\s*(\d{8,10})', texto_norm)
        for inicio, fin in rangos:
            try:
                for n in range(int(inicio), int(fin)+1):
                    radicados.add(str(n))
            except:
                pass
        texto_sin_rangos = re.sub(r'\d{8,10}\s*-\s*\d{8,10}', '', texto_norm)
        numeros = re.findall(r'\b(\d{8,10})\b', texto_sin_rangos)
        radicados.update(numeros)
        return sorted(radicados) if radicados else None
    
    def contar_paginas_tif(self, ruta):
        try:
            with PIL.Image.open(ruta) as img:
                return getattr(img, 'n_frames', 1)
        except:
            return 1
    
    def tamano_kb(self, ruta):
        try:
            return round(os.path.getsize(ruta)/1024, 1)
        except:
            return 0
    
    def extraer_radicado(self, ruta):
        partes = ruta.split(os.sep)
        for p in partes:
            if re.match(r'^\d{8,10}$', p):
                return p
        return "SIN_RADICADO"
    
    def buscar_archivos(self):
        carpeta = filedialog.askdirectory(title="Selecciona carpeta RAIZ")
        if not carpeta:
            return
        self.carpeta_raiz = carpeta
        self.log("="*70, "info")
        self.log(f"📂 Carpeta raíz: {carpeta}", "info")
        
        filtro = self.entry_filtro.get().strip().upper()
        if filtro:
            self.filtro_nombre = filtro
            self.lbl_filtro_activo.config(text=f"Filtro: '{filtro}'")
            self.log(f"🔍 Filtro nombre: '{filtro}'", "info")
        else:
            self.filtro_nombre = ""
            self.lbl_filtro_activo.config(text="Sin filtro")
        
        rad_texto = self.text_radicados.get("1.0", END).strip()
        rad_lista = self.parsear_radicados(rad_texto) if rad_texto else None
        if rad_lista:
            self.log(f"🔢 Buscando {len(rad_lista)} radicados específicos", "info")
        else:
            self.log("🔢 Sin radicados específicos, se escaneará toda la carpeta", "info")
        
        self.limpiar_lista()
        extensiones = {'.pdf','.tif','.tiff','.jpg','.jpeg','.png'}
        total_encontrados = 0
        total_paginas_tif = 0
        
        if rad_lista:
            for rad in rad_lista:
                ruta_rad = os.path.join(carpeta, rad)
                if not os.path.exists(ruta_rad):
                    self.log(f"❌ Radicado {rad}: carpeta no existe", "error")
                    continue
                archivo_encontrado = False
                for root, dirs, files in os.walk(ruta_rad):
                    for f in files:
                        ext = os.path.splitext(f)[1].lower()
                        if ext not in extensiones:
                            continue
                        if self.filtro_nombre and self.filtro_nombre.lower() not in f.lower():
                            continue
                        ruta_completa = os.path.join(root, f)
                        tipo = ext[1:].upper()
                        paginas = self.contar_paginas_tif(ruta_completa) if tipo in ['TIF','TIFF'] else 1
                        tamaño = self.tamano_kb(ruta_completa)
                        self.archivos_seleccionados.append({
                            'radicado': rad,
                            'ruta': ruta_completa,
                            'nombre': f,
                            'tipo': tipo,
                            'paginas': paginas,
                            'tamaño': tamaño
                        })
                        self.tree.insert("", END, text=str(len(self.archivos_seleccionados)),
                                         values=(rad, f, tipo, paginas, tamaño, ruta_completa))
                        total_encontrados += 1
                        if tipo in ['TIF','TIFF']:
                            total_paginas_tif += paginas
                        archivo_encontrado = True
                        break
                    if archivo_encontrado:
                        break
                if not archivo_encontrado:
                    self.log(f"⚠️ Radicado {rad}: no se encontraron archivos", "advertencia")
        else:
            for root, dirs, files in os.walk(carpeta):
                for f in files:
                    ext = os.path.splitext(f)[1].lower()
                    if ext not in extensiones:
                        continue
                    if self.filtro_nombre and self.filtro_nombre.lower() not in f.lower():
                        continue
                    ruta_completa = os.path.join(root, f)
                    rad = self.extraer_radicado(ruta_completa)
                    tipo = ext[1:].upper()
                    paginas = self.contar_paginas_tif(ruta_completa) if tipo in ['TIF','TIFF'] else 1
                    tamaño = self.tamano_kb(ruta_completa)
                    self.archivos_seleccionados.append({
                        'radicado': rad,
                        'ruta': ruta_completa,
                        'nombre': f,
                        'tipo': tipo,
                        'paginas': paginas,
                        'tamaño': tamaño
                    })
                    self.tree.insert("", END, text=str(len(self.archivos_seleccionados)),
                                     values=(rad, f, tipo, paginas, tamaño, ruta_completa))
                    total_encontrados += 1
                    if tipo in ['TIF','TIFF']:
                        total_paginas_tif += paginas
                    if total_encontrados <= 30:
                        self.log(f"   📄 {rad} → {f} ({tipo}, {paginas} págs, {tamaño}KB)", "info")
        
        if total_encontrados == 0:
            self.log("⚠️ No se encontraron archivos", "advertencia")
            messagebox.showwarning("Sin archivos", "No se encontraron archivos con los criterios dados")
        else:
            self.log(f"✅ Encontrados {total_encontrados} archivos", "exito")
            if total_paginas_tif:
                self.log(f"📄 Total páginas TIF/TIFF: {total_paginas_tif}", "info")
            messagebox.showinfo("Archivos encontrados", f"Se encontraron {total_encontrados} archivos")
    
    def limpiar_lista(self):
        self.archivos_seleccionados = []
        self.resultados = []
        for item in self.tree.get_children():
            self.tree.delete(item)
        self.btn_exportar.config(state=DISABLED)
        self.progress['value'] = 0
        self.lbl_progreso.config(text="Listo")
        self.lbl_tiempo.config(text="")
        self.log("🗑️ Lista limpiada", "advertencia")
    
    # Funciones de procesamiento de imágenes y OCR
    def mejorar_imagen(self, ruta_original):
        try:
            img = Image.open(ruta_original)
            if img.mode != 'L':
                img = img.convert('L')
            enhancer = ImageEnhance.Contrast(img)
            img = enhancer.enhance(2.5)
            img = img.filter(ImageFilter.UnsharpMask(radius=2, percent=150, threshold=3))
            img = ImageOps.equalize(img)
            arr = np.array(img)
            umbral = np.mean(arr) * 0.8
            arr = np.where(arr > umbral, 255, 0).astype(np.uint8)
            img = Image.fromarray(arr)
            temp_path = ruta_original.replace('.png', '_enh.png')
            if temp_path == ruta_original:
                temp_path = ruta_original + '_enh.png'
            img.save(temp_path, 'PNG')
            return temp_path
        except Exception as e:
            self.log(f"Error mejorando imagen: {e}", "error")
            return ruta_original
    
    def procesar_imagen(self, ruta_imagen, mejorar):
        try:
            if mejorar:
                ruta_proc = self.mejorar_imagen(ruta_imagen)
            else:
                ruta_proc = ruta_imagen
            try:
                settings = ocr.RecognitionSettings()
                settings.language = ocr.Language.SPA
                settings.auto_denoising = True
                settings.auto_contrast = True
                settings.auto_skew = True
                resultado = self.api.recognize(ruta_proc, settings)
            except AttributeError:
                inp = ocr.OcrInput(ocr.InputType.SINGLE_IMAGE)
                inp.add(ruta_proc)
                resultado = self.api.recognize(inp)
            if mejorar and ruta_proc != ruta_imagen and os.path.exists(ruta_proc):
                os.unlink(ruta_proc)
            if resultado and len(resultado) > 0:
                texto = resultado[0].recognition_text if hasattr(resultado[0], 'recognition_text') else str(resultado[0])
                texto = re.sub(r'[^\w\s\.\,\-\áéíóúñÑ\:\(\)]', ' ', texto)
                texto = re.sub(r'\s+', ' ', texto).strip()
                return texto
            return ""
        except Exception as e:
            self.log(f"      Error OCR: {e}", "error")
            return ""
    
    def procesar_pdf(self, ruta, dpi, mejorar):
        texto_total = []
        temp_files = []
        try:
            pages = convert_from_path(ruta, dpi=dpi, poppler_path=POPPLER_PATH)
            self.log(f"      📄 PDF con {len(pages)} páginas", "info")
            for i, pag in enumerate(pages):
                tmp = tempfile.NamedTemporaryFile(suffix='.png', delete=False)
                tmp_path = tmp.name
                tmp.close()
                pag.save(tmp_path, 'PNG')
                temp_files.append(tmp_path)
                txt = self.procesar_imagen(tmp_path, mejorar)
                if txt:
                    texto_total.append(txt)
                    self.log(f"      Página {i+1}: {len(txt)} caracteres", "info")
                else:
                    self.log(f"      Página {i+1}: sin texto", "advertencia")
            return ' '.join(texto_total), True
        except Exception as e:
            self.log(f"      Error PDF: {e}", "error")
            return "", False
        finally:
            for tmp in temp_files:
                try:
                    os.unlink(tmp)
                except:
                    pass
    
    def procesar_tif(self, ruta, mejorar):
        texto_total = []
        temp_files = []
        try:
            with PIL.Image.open(ruta) as img:
                n_pag = getattr(img, 'n_frames', 1)
                self.log(f"      🖼️ TIF con {n_pag} páginas", "info")
                for i in range(n_pag):
                    img.seek(i)
                    tmp = tempfile.NamedTemporaryFile(suffix='.png', delete=False)
                    tmp_path = tmp.name
                    tmp.close()
                    img.save(tmp_path, 'PNG')
                    temp_files.append(tmp_path)
                    txt = self.procesar_imagen(tmp_path, mejorar)
                    if txt:
                        texto_total.append(txt)
                        self.log(f"      Página {i+1}: {len(txt)} caracteres", "info")
                    else:
                        self.log(f"      Página {i+1}: sin texto", "advertencia")
            return ' '.join(texto_total), True
        except Exception as e:
            self.log(f"      Error TIF: {e}", "error")
            return "", False
        finally:
            for tmp in temp_files:
                try:
                    os.unlink(tmp)
                except:
                    pass
    
    def procesar_imagen_simple(self, ruta, mejorar):
        try:
            txt = self.procesar_imagen(ruta, mejorar)
            if txt:
                self.log(f"      📝 {len(txt)} caracteres extraídos", "info")
            else:
                self.log(f"      ⚠️ Sin texto detectado", "advertencia")
            return txt, True
        except Exception as e:
            self.log(f"      Error: {e}", "error")
            return "", False
    
    # ------------------------------------------------------------
    # PROCESAMIENTO PRINCIPAL
    # ------------------------------------------------------------
    def procesar_archivos(self):
        if not self.archivos_seleccionados:
            messagebox.showwarning("Sin archivos", "No hay archivos seleccionados")
            return
        if not self.api:
            messagebox.showerror("OCR", "OCR no inicializado")
            return
        
        palabras_input = self.entry_mis_palabras.get().strip()
        if not palabras_input:
            messagebox.showwarning("Sin palabras", "Ingresa al menos una palabra clave")
            return
        
        palabras_lista = [p.strip() for p in palabras_input.split(',') if p.strip()]
        tipo = self.tipo_busqueda.get()
        usar_fuzzy = self.fuzzy_var.get()
        try:
            longitud = int(self.entry_longitud.get())
            if longitud <= 0:
                longitud = LONGITUD_VALOR
        except:
            longitud = LONGITUD_VALOR
        
        dpi = int(self.combo_dpi.get())
        mejorar = self.mejorar_var.get()
        
        self.procesando = True
        self.cancelar = False
        self.resultados = []
        self.btn_procesar.config(state=DISABLED, text="⏳ EXTRAYENDO...")
        self.btn_cancelar.config(state=NORMAL)
        self.btn_exportar.config(state=DISABLED)
        self.progress['maximum'] = len(self.archivos_seleccionados)
        self.progress['value'] = 0
        
        self.log("="*70, "info")
        self.log("🚀 INICIANDO EXTRACCIÓN DE CAMPOS", "exito")
        self.log(f"📁 Archivos: {len(self.archivos_seleccionados)}", "info")
        self.log(f"🔑 Palabras: {palabras_lista}", "info")
        self.log(f"🎯 Tipo: {'TODAS' if tipo=='todas' else 'CUALQUIER'}", "info")
        self.log(f"🔍 Fuzzy: {'Sí' if usar_fuzzy else 'No'}", "info")
        self.log("="*70, "info")
        
        hilo = threading.Thread(target=self._hilo_extraer, args=(palabras_lista, tipo, usar_fuzzy, longitud, dpi, mejorar))
        hilo.daemon = True
        hilo.start()
    
    def _hilo_extraer(self, palabras, tipo, usar_fuzzy, longitud, dpi, mejorar):
        total = len(self.archivos_seleccionados)
        inicio = time.time()
        resultados_temp = []
        
        for i, arch in enumerate(self.archivos_seleccionados):
            if self.cancelar:
                self.root.after(0, lambda: self.log("⏹️ Proceso cancelado", "advertencia"))
                break
            
            self.root.after(0, lambda i=i, arch=arch: self.log(
                f"\n📄 [{i+1}/{total}] {arch['tipo']} | Rad: {arch['radicado']} | {arch['nombre'][:50]}", "info"))
            
            tiempo_archivo = time.time()
            
            if arch['tipo'] == 'PDF':
                texto_ocr, ok = self.procesar_pdf(arch['ruta'], dpi, mejorar)
            elif arch['tipo'] in ['TIF','TIFF']:
                texto_ocr, ok = self.procesar_tif(arch['ruta'], mejorar)
            else:
                texto_ocr, ok = self.procesar_imagen_simple(arch['ruta'], mejorar)
            
            duracion = time.time() - tiempo_archivo
            
            if not ok or not texto_ocr:
                resultados_por_palabra = {p: "ERROR OCR" for p in palabras}
                resultado = {
                    'radicado': arch['radicado'],
                    'archivo': arch['nombre'],
                    'ruta': arch['ruta'],
                    'tipo': arch['tipo'],
                    'paginas': arch['paginas'],
                    'tamaño': arch['tamaño'],
                    'encontrado': False,
                    'palabras_encontradas': [],
                    'resultados_por_palabra': resultados_por_palabra,
                    'texto_completo': "",
                    'tiempo': duracion
                }
                self.root.after(0, lambda: self.log("   ❌ Error al procesar", "error"))
            else:
                texto_norm = self.normalizar_texto(texto_ocr)
                resultados_por_palabra = {}
                palabras_encontradas = []
                
                for p in palabras:
                    valor = self.extraer_valor(texto_norm, p, longitud, usar_fuzzy)
                    if valor:
                        resultados_por_palabra[p] = valor
                        palabras_encontradas.append(p)
                        self.root.after(0, lambda p=p, v=valor[:80]: self.log(f"   ✅ '{p}' → {v}...", "exito"))
                    else:
                        resultados_por_palabra[p] = "NO ENCONTRADO"
                
                if tipo == "todas":
                    encontrado = len(palabras_encontradas) == len(palabras)
                else:
                    encontrado = len(palabras_encontradas) > 0
                
                resultado = {
                    'radicado': arch['radicado'],
                    'archivo': arch['nombre'],
                    'ruta': arch['ruta'],
                    'tipo': arch['tipo'],
                    'paginas': arch['paginas'],
                    'tamaño': arch['tamaño'],
                    'encontrado': encontrado,
                    'palabras_encontradas': palabras_encontradas,
                    'resultados_por_palabra': resultados_por_palabra,
                    'texto_completo': texto_ocr[:500] + ("..." if len(texto_ocr)>500 else ""),
                    'tiempo': duracion
                }
                if not encontrado:
                    self.root.after(0, lambda: self.log("   ❌ No se cumplió la condición", "error"))
            
            resultados_temp.append(resultado)
            self.root.after(0, self.progress.step)
            transcurrido = time.time() - inicio
            vel = (i+1) / transcurrido if transcurrido>0 else 0
            restantes = total - (i+1)
            tiempo_rest = restantes / vel if vel>0 else 0
            self.root.after(0, lambda i=i, tr=tiempo_rest: self._actualizar_progreso(i+1, total, tr))
        
        self.resultados = resultados_temp
        total_tiempo = time.time() - inicio
        self.root.after(0, lambda: self.log("="*70, "info"))
        self.root.after(0, lambda: self.log("✅ EXTRACCIÓN COMPLETADA", "exito"))
        self.root.after(0, lambda: self.log(f"⏱️ Tiempo total: {self._formatear_tiempo(total_tiempo)}", "info"))
        self.root.after(0, self._finalizar_procesamiento)
    
    def _actualizar_progreso(self, actual, total, tiempo_rest):
        pct = actual/total*100
        self.lbl_progreso.config(text=f"Progreso: {actual}/{total} ({pct:.1f}%)")
        self.lbl_tiempo.config(text=f"Tiempo restante: ~{self._formatear_tiempo(tiempo_rest)}")
    
    def _formatear_tiempo(self, seg):
        if seg < 60:
            return f"{seg:.0f} seg"
        elif seg < 3600:
            return f"{int(seg//60)} min {int(seg%60)} seg"
        else:
            return f"{int(seg//3600)} h {int((seg%3600)//60)} min"
    
    def cancelar_proceso(self):
        self.cancelar = True
        self.log("Cancelando proceso...", "advertencia")
    
    def _finalizar_procesamiento(self):
        self.procesando = False
        self.btn_procesar.config(state=NORMAL, text="▶️ EXTRAER CAMPOS")
        self.btn_cancelar.config(state=DISABLED)
        self.btn_exportar.config(state=NORMAL)
        self.lbl_progreso.config(text="¡Completado!")
        self.lbl_tiempo.config(text="")
        encontrados = sum(1 for r in self.resultados if r['encontrado'])
        messagebox.showinfo("Completado", f"Extracción finalizada\n\nTotal: {len(self.resultados)}\nCoincidencias: {encontrados}")
    
    # ------------------------------------------------------------
    # EXPORTAR A EXCEL
    # ------------------------------------------------------------
    def exportar_excel(self):
        if not self.resultados:
            messagebox.showwarning("Sin resultados", "No hay resultados para exportar")
            return
        archivo = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"extraccion_campos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        if not archivo:
            return
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Campos Extraídos"
            
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            encontrado_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            
            palabras_clave = list(self.resultados[0]['resultados_por_palabra'].keys()) if self.resultados else []
            headers = ["Radicado", "Archivo", "Tipo", "Páginas", "Tamaño(KB)", "Ruta"] + palabras_clave + ["Estado", "Tiempo(s)"]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            for i, r in enumerate(self.resultados, 2):
                ws.cell(row=i, column=1, value=r['radicado'])
                ws.cell(row=i, column=2, value=r['archivo'])
                ws.cell(row=i, column=3, value=r['tipo'])
                ws.cell(row=i, column=4, value=r['paginas'])
                ws.cell(row=i, column=5, value=r['tamaño'])
                ws.cell(row=i, column=6, value=r['ruta'])
                col_offset = 6
                for idx, p in enumerate(palabras_clave):
                    valor = r['resultados_por_palabra'].get(p, "")
                    ws.cell(row=i, column=col_offset + idx + 1, value=valor)
                    if valor not in ("NO ENCONTRADO", "ERROR OCR"):
                        ws.cell(row=i, column=col_offset + idx + 1).fill = encontrado_fill
                estado = "✅ ENCONTRADO" if r['encontrado'] else "❌ NO ENCONTRADO"
                ws.cell(row=i, column=col_offset + len(palabras_clave) + 1, value=estado)
                ws.cell(row=i, column=col_offset + len(palabras_clave) + 2, value=round(r['tiempo'],2))
            
            anchos = [15, 60, 10, 10, 12, 80] + [35]*len(palabras_clave) + [20, 10]
            for i, ancho in enumerate(anchos, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = ancho
            
            wb.save(archivo)
            self.log(f"📊 Excel guardado: {archivo}", "exito")
            messagebox.showinfo("Exportado", f"Excel guardado en:\n{archivo}")
        except Exception as e:
            self.log(f"Error exportando Excel: {e}", "error")
            messagebox.showerror("Error", f"No se pudo exportar: {e}")
    
    # ------------------------------------------------------------
    # PRUEBA MANUAL
    # ------------------------------------------------------------
    def probar_ocr_manual(self):
        archivo = filedialog.askopenfilename(title="Selecciona un archivo para probar",
            filetypes=[("PDF","*.pdf"),("Imágenes","*.tif *.tiff *.jpg *.png"),("Todos","*.*")])
        if not archivo:
            return
        self.log("="*70, "diagnostico")
        self.log(f"🧪 PRUEBA OCR: {os.path.basename(archivo)}", "diagnostico")
        ext = os.path.splitext(archivo)[1].lower()
        dpi = int(self.combo_dpi.get())
        mejorar = self.mejorar_var.get()
        
        if ext == '.pdf':
            texto, ok = self.procesar_pdf(archivo, dpi, mejorar)
        elif ext in ['.tif','.tiff']:
            texto, ok = self.procesar_tif(archivo, mejorar)
        else:
            texto, ok = self.procesar_imagen_simple(archivo, mejorar)
        
        if not ok or not texto:
            self.log("❌ No se pudo extraer texto", "error")
            return
        
        self.log("✅ Texto extraído. Mostrando primeros 500 caracteres:", "exito")
        self.log(texto[:500], "info")
        
        # Probar extracción de campos con las palabras actuales
        palabras_input = self.entry_mis_palabras.get().strip()
        if palabras_input:
            palabras = [p.strip() for p in palabras_input.split(',') if p.strip()]
            texto_norm = self.normalizar_texto(texto)
            usar_fuzzy = self.fuzzy_var.get()
            longitud = int(self.entry_longitud.get()) if self.entry_longitud.get().isdigit() else LONGITUD_VALOR
            self.log("🔍 Probando extracción de campos:", "info")
            for p in palabras:
                valor = self.extraer_valor(texto_norm, p, longitud, usar_fuzzy)
                if valor:
                    self.log(f"   ✅ {p}: {valor}", "exito")
                else:
                    self.log(f"   ❌ {p}: NO ENCONTRADO", "error")

# ------------------------------------------------------------
# EJECUTAR
# ------------------------------------------------------------
if __name__ == "__main__":
    root = Tk()
    app = AppOCR(root)
    root.mainloop()